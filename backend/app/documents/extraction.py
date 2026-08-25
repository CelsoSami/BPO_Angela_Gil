"""Extração de informações estruturadas de documentos.

Fase 1: engine `regex` (PDF/XLSX/CSV) — sem IA.
Fase 2: a engine pode ser trocada por OCR/LLM mantendo a mesma interface
(`extract_fields(texto, tipo_documento) -> dict[campo, valor]`).
"""
import csv
import io
import re
from datetime import date

from app.config import settings

CAMPOS_PADRAO = [
    "document_type",
    "document_number",
    "issue_date",
    "due_date",
    "supplier",
    "amount",
]

_EXTRAIDO = object()


def extract_text(arquivo_path: str, mime_type: str | None = None) -> str:
    """Extrai o texto bruto do arquivo (PDF, XLSX, CSV). Imagens: ver nota."""
    ext = (arquivo_path or "").rsplit(".", 1)[-1].lower() if arquivo_path else ""
    if ext == "pdf" or (mime_type and "pdf" in mime_type):
        from pypdf import PdfReader

        reader = PdfReader(arquivo_path)
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    if ext in ("xlsx", "xls") or (mime_type and "spreadsheet" in mime_type):
        from openpyxl import load_workbook

        wb = load_workbook(arquivo_path, read_only=True, data_only=True)
        partes = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                partes.append(" | ".join(str(c) for c in row if c is not None))
        return "\n".join(partes)
    if ext == "csv" or (mime_type and "csv" in mime_type):
        with open(arquivo_path, "r", encoding="utf-8-sig", errors="replace") as fh:
            return fh.read()
    # Imagens (PNG/JPG): OCR não faz parte da Fase 1.
    # Integração futura: engine OCR aqui (ex.: Tesseract/IA).
    raise ValueError("OCR para imagens estará disponível na Fase 2.")


def _find_amount(texto: str) -> str | None:
    padroes = [
        r"Total[:\s]*R\$\s*([\d.,]+)",
        r"Valor (?:total|a pagar|da nota)[:\s]*R\$\s*([\d.,]+)",
        r"R\$\s*([\d]{1,3}(?:\.\d{3})*(?:,\d{2})?)",
    ]
    for p in padroes:
        m = re.search(p, texto, re.IGNORECASE)
        if m:
            return m.group(1)
    return None


def _normalize_amount(raw: str | None) -> str | None:
    if not raw:
        return None
    s = raw.strip().replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        # sem centavos: o ponto é separador de milhar (ex.: 1.234)
        if re.fullmatch(r"\d{1,3}(?:\.\d{3})+", s):
            s = s.replace(".", "")
    return s


def _find_dates(texto: str) -> list[str]:
    return re.findall(r"\b(\d{2})[/\-.](\d{2})[/\-.](\d{4})\b", texto)


def _find_supplier(texto: str) -> str | None:
    for label in ("Fornecedor", "Emitente", "Prestador", "Razão Social", "Cliente"):
        m = re.search(
            rf"{label}\s*[:.]?\s*([A-Za-zÀ-ú0-9][^\n|]{{3,80}})", texto, re.IGNORECASE
        )
        if m:
            return m.group(1).strip()
    return None


def _find_number(texto: str) -> str | None:
    m = re.search(r"\b(CT[-/\s]?\d{2,4}[-/\s]?\d{3,4})\b", texto, re.IGNORECASE)
    if m:
        return m.group(1).replace(" ", "")
    m = re.search(r"\b(NF[-e]?\s*\d{3,9})\b", texto, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"\b(Recibo\s*(?:n[ºo]?\.?\s*)?\d{2,8})\b", texto, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


def _find_document_type(texto: str) -> str | None:
    mapa = [
        (r"contrato", "CONTRATO"),
        (r"nota fiscal", "NOTA_FISCAL"),
        (r"recibo", "RECIBO"),
        (r"comprovante", "COMPROVANTE"),
        (r"extrato", "EXTRATO"),
    ]
    for padrao, tipo in mapa:
        if re.search(padrao, texto, re.IGNORECASE):
            return tipo
    return "OUTRO"


def extract_fields(texto: str, tipo_documento: str | None = None) -> dict:
    """Extrai os campos estruturados do texto. Retorna dict campo -> valor."""
    doc_type = _find_document_type(texto) if not tipo_documento else tipo_documento
    datas = _find_dates(texto)
    amount_raw = _find_amount(texto)

    campos = {
        "document_type": doc_type,
        "document_number": _find_number(texto),
        "issue_date": datas[0] if datas else None,
        "due_date": datas[-1] if len(datas) > 1 else None,
        "supplier": _find_supplier(texto),
        "amount": _normalize_amount(amount_raw),
    }
    return {k: v for k, v in campos.items() if v}


def process_document_file(db, documento, arquivo_path: str) -> int:
    """Processa um documento salvo: extrai campos e grava em document_extractions.

    Retorna quantos campos foram extraídos. O documento passa a
    AGUARDANDO_VALIDACAO quando há campos a validar, senão PROCESSADO.
    """
    from app.models.documents import DocumentExtraction

    try:
        texto = extract_text(arquivo_path, documento.mime_type)
    except Exception:
        texto = ""
    if not texto.strip():
        documento.status = "PROCESSADO"
        db.flush()
        return 0

    campos = extract_fields(texto, documento.tipo if documento.tipo != "OUTRO" else None)
    count = 0
    for campo, valor in campos.items():
        valor_str = str(valor)
        if campo == "issue_date" or campo == "due_date":
            # normaliza dd/mm/aaaa -> aaaa-mm-dd
            m = re.match(r"(\d{2})[/\-.](\d{2})[/\-.](\d{4})", valor_str)
            if m:
                valor_str = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        existente = (
            db.query(DocumentExtraction)
            .filter(
                DocumentExtraction.document_id == documento.id,
                DocumentExtraction.campo == campo,
            )
            .first()
        )
        if existente:
            existente.valor = valor_str
            existente.status = "EXTRAIDA"
        else:
            db.add(
                DocumentExtraction(
                    document_id=documento.id,
                    campo=campo,
                    valor=valor_str,
                    status="EXTRAIDA",
                )
            )
        count += 1

    documento.status = "AGUARDANDO_VALIDACAO" if count else "PROCESSADO"
    db.flush()
    return count
