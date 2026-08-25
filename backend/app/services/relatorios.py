"""Geração do Relatório Gerencial Mensal (baseado em dados — sem IA)."""
import csv
import io
from datetime import date

from sqlalchemy.orm import Session

from app.models.admin import Report
from app.models.alerts import Alert
from app.models.clients import Client
from app.models.contracts import Contract
from app.models.documents import Document
from app.models.financial import Payable, Receivable
from app.models.projects import Project
from app.services import inadimplencia
from app.services.cashflow import summary, trend
from app.services.dre import compute_dre
from app.services.indicators import compute_client_kpis
from app.services.rentabilidade import compute_profitability
from app.utils.dates import month_range

MESES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def _money(v) -> str:
    return f"R$ {float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _pct(v) -> str:
    return f"{float(v or 0):.1f}%"


def build_monthly_report(db: Session, client_id, mes: int, ano: int) -> dict:
    """Monta o conteúdo completo do relatório gerencial mensal (13 seções)."""
    client = db.get(Client, client_id)
    if client is None:
        raise ValueError("Cliente não encontrado.")
    start, end = month_range(mes, ano)

    dre = compute_dre(db, client_id, mes, ano)
    fc = summary(db, client_id, start, end)
    tendencia = trend(db, client_id)
    inad = inadimplencia.compute_panel(db, client_id=client_id)
    kpis = compute_client_kpis(db, client_id)

    receber = (
        db.query(Receivable)
        .filter(Receivable.client_id == client_id, Receivable.status != "RECEBIDO")
        .order_by(Receivable.vencimento)
        .all()
    )
    pagar = (
        db.query(Payable)
        .filter(Payable.client_id == client_id, Payable.status != "PAGO")
        .order_by(Payable.vencimento)
        .all()
    )
    projetos = db.query(Project).filter(Project.client_id == client_id).all()
    rent = compute_profitability(db, client_id=client_id)
    contratos = (
        db.query(Contract)
        .filter(Contract.client_id == client_id, Contract.status.in_(["ATIVO", "EM_ANALISE", "PENDENTE"]))
        .all()
    )
    docs = (
        db.query(Document)
        .filter(
            Document.client_id == client_id,
            Document.status.in_(["PENDENTE", "AGUARDANDO_VALIDACAO"]),
        )
        .all()
    )
    alertas = (
        db.query(Alert)
        .filter(Alert.client_id == client_id, Alert.status.in_(["ABERTO", "EM_ANDAMENTO"]))
        .order_by(Alert.prioridade)
        .all()
    )

    pontos_atencao = [a.titulo for a in alertas if a.prioridade == "ALTA"] or [
        "Nenhum ponto crítico identificado."
    ]

    resumo = (
        f"No mês de {MESES[mes-1]}/{ano}, a receita foi de {_money(dre['receita_bruta'])} "
        f"e as despesas de {_money(fc['saidas'])}, resultando em saldo de "
        f"{_money(fc['saldo_final'])}. O resultado líquido (DRE) foi "
        f"{_money(dre['resultado_liquido'])} ({_pct(dre['resultado_liquido'] / dre['receita_bruta'] * 100) if dre['receita_bruta'] else '—'} da receita). "
        f"Indicador de saúde financeira: {kpis.get('saude_financeira', {}).get('classificacao', '—')}."
    )

    return {
        "cliente": client.nome_fantasia or client.razao_social,
        "mes": mes,
        "ano": ano,
        "gerado_em": date.today().isoformat(),
        "secoes": {
            "resumo_executivo": resumo,
            "resultado_financeiro": {
                "receita": dre["receita_bruta"],
                "despesas": fc["saidas"],
                "resultado": dre["resultado_liquido"],
                "margem": dre["resultado_liquido"] / dre["receita_bruta"] * 100
                if dre["receita_bruta"] else 0,
            },
            "fluxo_de_caixa": {
                "entradas": fc["entradas"],
                "saidas": fc["saidas"],
                "saldo_final": fc["saldo_final"],
                "tendencia": tendencia,
            },
            "dre": dre,
            "contas_a_receber": [
                {
                    "descricao": r.descricao or "—",
                    "vencimento": r.vencimento.isoformat() if r.vencimento else "—",
                    "valor": float(r.valor),
                    "status": r.status,
                }
                for r in receber
            ],
            "inadimplencia": inad,
            "contas_a_pagar": [
                {
                    "fornecedor": p.fornecedor,
                    "vencimento": p.vencimento.isoformat() if p.vencimento else "—",
                    "valor": float(p.valor),
                    "status": p.status,
                }
                for p in pagar
            ],
            "projetos": [
                {
                    "nome": p.nome,
                    "status": p.status,
                    "receita": float(p.receita or 0),
                    "custo_realizado": float(p.custo_realizado or 0),
                }
                for p in projetos
            ],
            "rentabilidade": rent,
            "indicadores": kpis,
            "contratos": [
                {
                    "numero": c.numero,
                    "valor": float(c.valor or 0),
                    "termino": c.termino.isoformat() if c.termino else "—",
                    "status": c.status,
                }
                for c in contratos
            ],
            "pendencias": {
                "documentos": [
                    {"arquivo": d.arquivo_nome, "status": d.status} for d in docs
                ],
                "alertas_abertos": [a.titulo for a in alertas],
            },
            "pontos_de_atencao": pontos_atencao,
        },
    }


def save_report_record(db: Session, client_id, mes, ano, conteudo: dict, user) -> Report:
    rep = Report(
        client_id=client_id,
        tipo="MENSAL",
        mes=mes,
        ano=ano,
        titulo=f"Relatório Gerencial Mensal — {conteudo.get('cliente', '')} ({mes:02d}/{ano})",
        conteudo_json=conteudo,
        criado_por=user.id if user else None,
    )
    db.add(rep)
    db.flush()
    return rep


# ------------------------------------------------------------------ EXPORTS
def export_csv(sections: dict) -> str:
    """Exporta as principais tabelas do relatório em CSV (arquivo único, seções)."""
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["BUILD FLOW BPO — RELATÓRIO GERENCIAL MENSAL"])
    writer.writerow([f"Cliente: {sections['cliente']}  |  Mês: {sections['mes']:02d}/{sections['ano']}"])
    writer.writerow([])
    writer.writerow(["RESULTADO FINANCEIRO"])
    r = sections["secoes"]["resultado_financeiro"]
    writer.writerow(["Receita", "Despesas", "Resultado", "Margem %"])
    writer.writerow([r["receita"], r["despesas"], r["resultado"], round(r["margem"], 1)])
    writer.writerow([])
    writer.writerow(["DRE"])
    dre = sections["secoes"]["dre"]
    for k in [
        "receita_bruta", "impostos", "receita_liquida", "custos_diretos",
        "margem_contribuicao", "despesas_operacionais", "resultado_operacional",
        "despesas_financeiras", "resultado_liquido",
    ]:
        writer.writerow([k, dre[k]])
    writer.writerow([])
    writer.writerow(["CONTAS A RECEBER"])
    writer.writerow(["Descrição", "Vencimento", "Valor", "Status"])
    for row in sections["secoes"]["contas_a_receber"]:
        writer.writerow([row["descricao"], row["vencimento"], row["valor"], row["status"]])
    writer.writerow([])
    writer.writerow(["CONTAS A PAGAR"])
    writer.writerow(["Fornecedor", "Vencimento", "Valor", "Status"])
    for row in sections["secoes"]["contas_a_pagar"]:
        writer.writerow([row["fornecedor"], row["vencimento"], row["valor"], row["status"]])
    writer.writerow([])
    writer.writerow(["RENTABILIDADE POR PROJETO"])
    writer.writerow(["Projeto", "Receita", "Custo Realizado", "Lucro", "Margem %"])
    for row in sections["secoes"]["rentabilidade"]:
        writer.writerow([row["nome"], row["receita"], row["custo_realizado"], row["lucro"], round(row["margem"], 1)])
    writer.writerow([])
    writer.writerow(["PONTOS DE ATENÇÃO"])
    for p in sections["secoes"]["pontos_de_atencao"]:
        writer.writerow([p])
    return out.getvalue()


def export_excel(sections: dict) -> bytes:
    """Exporta o relatório em .xlsx com abas por seção."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["BUILD FLOW BPO — RELATÓRIO GERENCIAL MENSAL"])
    ws.append([f"Cliente: {sections['cliente']}  |  Mês: {sections['mes']:02d}/{sections['ano']}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    r = sections["secoes"]["resultado_financeiro"]
    ws.append(["Receita", "Despesas", "Resultado", "Margem %"])
    ws.append([r["receita"], r["despesas"], r["resultado"], round(r["margem"], 1)])
    ws.append([])
    ws.append(["Resumo executivo"])
    ws.append([sections["secoes"]["resumo_executivo"]])

    def _tabela(nome, headers, rows):
        sh = wb.create_sheet(nome)
        sh.append(headers)
        for row in rows:
            sh.append(row)
        for cell in sh[1]:
            cell.font = Font(bold=True)

    dre = sections["secoes"]["dre"]
    _tabela("DRE", ["Conta", "Valor"], [[k, dre[k]] for k in dre if k not in ("mes", "ano")])
    _tabela("A Receber", ["Descrição", "Vencimento", "Valor", "Status"],
            [[x["descricao"], x["vencimento"], x["valor"], x["status"]] for x in sections["secoes"]["contas_a_receber"]])
    _tabela("A Pagar", ["Fornecedor", "Vencimento", "Valor", "Status"],
            [[x["fornecedor"], x["vencimento"], x["valor"], x["status"]] for x in sections["secoes"]["contas_a_pagar"]])
    _tabela("Projetos", ["Projeto", "Receita", "Custo Realizado", "Lucro", "Margem %"],
            [[x["nome"], x["receita"], x["custo_realizado"], x["lucro"], round(x["margem"], 1)] for x in sections["secoes"]["rentabilidade"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sanitize(text: str) -> str:
    """Torna o texto seguro para fontes Latin-1 do reportlab."""
    if not text:
        return ""
    repl = {
        "—": "-", "–": "-", "•": "-", "·": "-", "“": '"', "”": '"',
        "’": "'", "‘": "'", "…": "...", "×": "x", "²": "2", "³": "3",
        "⇒": "->", "→": "->",
    }
    for k, v in repl.items():
        text = text.replace(k, v)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def export_pdf(sections: dict) -> bytes:
    """Exporta o relatório em PDF (reportlab)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Relatório Mensal — {sections['cliente']}",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("TitleBF", parent=styles["Title"], fontSize=17, spaceAfter=4)
    h2 = ParagraphStyle("H2BF", parent=styles["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("BodyBF", parent=styles["BodyText"], fontSize=9, leading=12)

    story = [
        Paragraph("BUILD FLOW BPO", title),
        Paragraph(
            _sanitize(
                f"Relatório Gerencial Mensal — {sections['cliente']} "
                f"({sections['mes']:02d}/{sections['ano']}) — gerado em {sections['gerado_em']}"
            ),
            body,
        ),
        Spacer(1, 6),
    ]
    sec = sections["secoes"]

    def add_table(headers, rows):
        data = [[_sanitize(str(c)) for c in headers]] + [
            [_sanitize(str(c)) for c in row] for row in rows
        ]
        t = Table(data, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#101216")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F2F4")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(t)

    story.append(Paragraph("1. Resumo Executivo", h2))
    story.append(Paragraph(_sanitize(sec["resumo_executivo"]), body))

    story.append(Paragraph("2. Resultado Financeiro", h2))
    r = sec["resultado_financeiro"]
    add_table(["Indicador", "Valor"], [
        ["Receita", _money(r["receita"])],
        ["Despesas", _money(r["despesas"])],
        ["Resultado", _money(r["resultado"])],
        ["Margem", _pct(r["margem"])],
    ])

    story.append(Paragraph("3. Fluxo de Caixa", h2))
    fc = sec["fluxo_de_caixa"]
    add_table(["Entradas", "Saídas", "Saldo Final"], [
        [_money(fc["entradas"]), _money(fc["saidas"]), _money(fc["saldo_final"])],
    ])

    story.append(Paragraph("4. DRE", h2))
    dre = sec["dre"]
    add_table(["Conta", "Valor"], [
        ["Receita Bruta", _money(dre["receita_bruta"])],
        ["(-) Impostos", _money(dre["impostos"])],
        ["= Receita Líquida", _money(dre["receita_liquida"])],
        ["(-) Custos Diretos", _money(dre["custos_diretos"])],
        ["= Margem de Contribuição", _money(dre["margem_contribuicao"])],
        ["(-) Despesas Operacionais", _money(dre["despesas_operacionais"])],
        ["= Resultado Operacional", _money(dre["resultado_operacional"])],
        ["(-) Despesas Financeiras", _money(dre["despesas_financeiras"])],
        ["= Resultado Líquido", _money(dre["resultado_liquido"])],
    ])

    story.append(Paragraph("5. Contas a Receber", h2))
    add_table(["Descrição", "Vencimento", "Valor", "Status"], [
        [x["descricao"], x["vencimento"], _money(x["valor"]), x["status"]]
        for x in sec["contas_a_receber"]
    ])

    story.append(Paragraph("6. Inadimplência", h2))
    inad = sec["inadimplencia"]
    add_table(["Indicador", "Valor"], [
        ["Total vencido", _money(inad["total_vencido"])],
        ["Títulos vencidos", inad["quantidade_titulos"]],
        ["Dias médios de atraso", inad["dias_medio_atraso"]],
    ])
    add_table(["Faixa", "Títulos", "Valor"], [
        [f["faixa"], f["titulos"], _money(f["valor"])] for f in inad["faixas"]
    ])

    story.append(Paragraph("7. Contas a Pagar", h2))
    add_table(["Fornecedor", "Vencimento", "Valor", "Status"], [
        [x["fornecedor"], x["vencimento"], _money(x["valor"]), x["status"]]
        for x in sec["contas_a_pagar"]
    ])

    story.append(Paragraph("8. Projetos", h2))
    add_table(["Projeto", "Status", "Receita", "Custo Realizado"], [
        [x["nome"], x["status"], _money(x["receita"]), _money(x["custo_realizado"])]
        for x in sec["projetos"]
    ])

    story.append(Paragraph("9. Rentabilidade", h2))
    add_table(["Projeto", "Receita", "Custo", "Lucro", "Margem"], [
        [x["nome"], _money(x["receita"]), _money(x["custo_realizado"]), _money(x["lucro"]), _pct(x["margem"])]
        for x in sec["rentabilidade"]
    ])

    story.append(Paragraph("10. Indicadores", h2))
    k = sec["indicadores"]
    add_table(["Indicador", "Valor"], [
        ["Receita total", _money(k["receita"])],
        ["Despesas", _money(k["despesas"])],
        ["Margem", _pct(k["margem"])],
        ["A receber", _money(k["a_receber"])],
        ["Atrasado", _money(k["atrasado"])],
        ["Projetos ativos", k["projetos_ativos"]],
        ["Contratos ativos", k["contratos_ativos"]],
    ])

    story.append(Paragraph("11. Contratos", h2))
    add_table(["Número", "Valor", "Término", "Status"], [
        [x["numero"], _money(x["valor"]), x["termino"], x["status"]]
        for x in sec["contratos"]
    ])

    story.append(Paragraph("12. Pendências", h2))
    pend = sec["pendencias"]
    add_table(["Documentos pendentes", "Alertas abertos"], [
        [len(pend["documentos"]), len(pend["alertas_abertos"])],
    ])

    story.append(Paragraph("13. Pontos de Atenção", h2))
    for p in sec["pontos_de_atencao"]:
        story.append(Paragraph(f"- {_sanitize(p)}", body))

    doc.build(story)
    return buf.getvalue()
